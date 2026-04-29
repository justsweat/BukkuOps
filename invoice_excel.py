from datetime import datetime, timedelta
from openpyxl import load_workbook
import ops as o


excel_path = "data/bukku_invoices.xlsx"

def update_invoices(excel_path=excel_path):
    """Retrieve all invoices from the past 14 days and upsert them into the Excel file.

    The function paginates through the Bukku API, then for each invoice it either updates the
    existing amount (if the invoice number is already present) or appends a new row.
    """
    today = datetime.now().date()
    date_to = today.isoformat()
    date_from = (today - timedelta(days=14)).isoformat()

    path = "sales/invoices"

    wb = load_workbook(excel_path)
    ws = wb["Bukku"]

    # Build a lookup of existing invoice numbers to their row index (assuming number is in column A)
    existing = {}
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        cell = row[0]
        if cell.value:
            existing[cell.value] = cell.row

    added = 0
    updated = 0
    page = 1
    while True:
        params = {
            "date_from": date_from,
            "date_to": date_to,
            "page_size": 30,
            "page": page,
        }
        response = o.get_response(path, params)
        transactions = response.get("transactions", [])
        if not transactions:
            break
        for txn in transactions:
            number = txn.get("number")
            amount = txn.get("amount")
            if number in existing:
                ws.cell(row=existing[number], column=2, value=amount)
                updated += 1
            else:
                ws.append([number, amount])
                existing[number] = ws.max_row
                added += 1
        paging = response.get("paging", {})
        total_pages = paging.get("total", 1)
        if page >= total_pages:
            break
        page += 1

    wb.save(excel_path)
    print(f"Invoices added: {added}, updated: {updated}")


if __name__ == "__main__":
    update_invoices()
